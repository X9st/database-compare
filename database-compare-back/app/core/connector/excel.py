"""Excel 文件连接器"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .base import BaseConnector, TableInfo, ColumnInfo, IndexInfo, ConstraintInfo


class ExcelConnector(BaseConnector):
    """Excel 文件连接器，支持 xlsx/xls。"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.source_kind = "file"
        self._mode = "single_file"
        self._workbook = None
        self._storage_path: Optional[Path] = None
        self._file_type: Optional[str] = None
        self._table_index: Dict[str, Dict[str, Any]] = {}

    def _resolve_mode(self) -> str:
        mode = str((self.extra_config or {}).get("mode") or "").strip().lower()
        if mode:
            return mode
        snapshot = (self.extra_config or {}).get("snapshot")
        if isinstance(snapshot, dict) and isinstance(snapshot.get("table_index"), dict):
            return "remote_dataset"
        return "single_file"

    def _resolve_storage_path(self) -> Path:
        storage_key = str((self.extra_config or {}).get("storage_key") or "").strip()
        if not storage_key:
            raise ConnectionError("Excel 数据源缺少 extra_config.storage_key")
        path = Path(storage_key)
        if not path.is_absolute():
            path = Path.cwd() / path
        return path.resolve()

    def _header_row(self) -> int:
        raw = (self.extra_config or {}).get("header_row", 1)
        try:
            value = int(raw)
            return value if value >= 1 else 1
        except Exception:
            return 1

    def connect(self) -> bool:
        self._mode = self._resolve_mode()
        if self._mode == "remote_dataset":
            snapshot = (self.extra_config or {}).get("snapshot") or {}
            table_index = snapshot.get("table_index")
            if not isinstance(table_index, dict) or not table_index:
                raise ConnectionError("Excel 远程数据集缺少 snapshot.table_index")
            normalized: Dict[str, Dict[str, Any]] = {}
            for table_name, entry in table_index.items():
                if not isinstance(entry, dict):
                    continue
                file_path = Path(str(entry.get("storage_key") or ""))
                if not file_path.is_absolute():
                    file_path = Path.cwd() / file_path
                file_path = file_path.resolve()
                if not file_path.exists() or not file_path.is_file():
                    raise ConnectionError(f"Excel 数据集文件不存在: {file_path}")
                normalized[str(table_name)] = {**entry, "storage_key": str(file_path.as_posix())}
            if not normalized:
                raise ConnectionError("Excel 远程数据集无可用表")
            self._table_index = normalized
            return True

        path = self._resolve_storage_path()
        if not path.exists() or not path.is_file():
            raise ConnectionError(f"Excel 文件不存在: {path}")
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".xls"}:
            raise ConnectionError("Excel 文件仅支持 .xlsx/.xls")

        self._storage_path = path
        self._file_type = suffix.lstrip(".")
        try:
            if suffix == ".xlsx":
                from openpyxl import load_workbook

                self._workbook = load_workbook(path, data_only=True, read_only=True)
            else:
                import xlrd

                self._workbook = xlrd.open_workbook(path)
            return True
        except ImportError as exc:
            raise ConnectionError(f"Excel 驱动缺失: {exc}") from exc
        except Exception as exc:
            raise ConnectionError(f"Excel 打开失败: {exc}") from exc

    def disconnect(self) -> None:
        if self._workbook is not None and hasattr(self._workbook, "close"):
            try:
                self._workbook.close()
            except Exception:
                pass
        self._workbook = None
        self._table_index = {}

    def test_connection(self) -> Dict[str, Any]:
        try:
            self.connect()
            version = self.get_version()
            self.disconnect()
            return {"success": True, "message": "连接成功", "version": version}
        except Exception as exc:
            return {"success": False, "message": str(exc), "version": None}

    def _sheet_names(self) -> List[str]:
        if self._workbook is None:
            raise ConnectionError("Excel 未连接")
        if self._file_type == "xlsx":
            return list(self._workbook.sheetnames)
        return list(self._workbook.sheet_names())

    def _resolve_dataset_entry(self, table_name: str) -> Dict[str, Any]:
        target = str(table_name or "").strip().lower()
        for name, entry in self._table_index.items():
            if name.lower() == target:
                return entry
        raise ValueError(f"表不存在: {table_name}")

    def _resolve_sheet_name(self, table_name: str) -> str:
        target = str(table_name or "").strip().lower()
        for name in self._sheet_names():
            if name.lower() == target:
                return name
        raise ValueError(f"Sheet 不存在: {table_name}")

    def _read_rows_from_file(
        self,
        path: Path,
        file_type: str,
        sheet_name: Optional[str],
    ) -> Tuple[List[str], List[Dict[str, Any]]]:
        header_row = self._header_row()
        if file_type == "xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, data_only=True, read_only=True)
            try:
                target_sheet = sheet_name or workbook.sheetnames[0]
                sheet = workbook[target_sheet]
                raw_rows = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()
        elif file_type == "xls":
            import xlrd

            workbook = xlrd.open_workbook(path)
            target_sheet = sheet_name or workbook.sheet_names()[0]
            sheet = workbook.sheet_by_name(target_sheet)
            raw_rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        else:
            raise ValueError(f"不支持的 Excel 类型: {file_type}")

        if not raw_rows:
            return [], []

        header_idx = min(header_row - 1, max(len(raw_rows) - 1, 0))
        headers = self._normalize_headers(list(raw_rows[header_idx]))
        data_rows = raw_rows[header_idx + 1 :]
        records: List[Dict[str, Any]] = []
        for row in data_rows:
            if len(row) < len(headers):
                row = list(row) + [None] * (len(headers) - len(row))
            values = row[: len(headers)]
            records.append(dict(zip(headers, values)))
        return headers, records

    def get_tables(self) -> List[TableInfo]:
        if self._mode == "remote_dataset":
            tables: List[TableInfo] = []
            for table_name, entry in self._table_index.items():
                storage_path = Path(str(entry.get("storage_key")))
                file_type = str(entry.get("file_type") or storage_path.suffix.lstrip(".")).lower()
                sheet_name = str(entry.get("sheet_name") or "").strip() or None
                _, rows = self._read_rows_from_file(storage_path, file_type, sheet_name)
                tables.append(
                    TableInfo(
                        name=table_name,
                        schema=str(entry.get("original_name") or storage_path.name),
                        comment=f"Excel Sheet: {sheet_name or table_name}",
                        row_count=len(rows),
                    )
                )
            return tables

        header_row = self._header_row()
        tables: List[TableInfo] = []
        for sheet_name in self._sheet_names():
            row_count = 0
            if self._file_type == "xlsx":
                sheet = self._workbook[sheet_name]
                row_count = max(sheet.max_row - header_row, 0)
            else:
                sheet = self._workbook.sheet_by_name(sheet_name)
                row_count = max(sheet.nrows - header_row, 0)
            tables.append(
                TableInfo(
                    name=sheet_name,
                    schema=self._storage_path.name if self._storage_path else None,
                    comment=f"Excel Sheet: {sheet_name}",
                    row_count=row_count,
                )
            )
        return tables

    def _normalize_headers(self, headers: List[Any]) -> List[str]:
        seen: Dict[str, int] = {}
        normalized: List[str] = []
        for idx, raw in enumerate(headers, start=1):
            text = str(raw).strip() if raw is not None else ""
            if not text:
                text = f"column_{idx}"
            lower = text.lower()
            seen[lower] = seen.get(lower, 0) + 1
            if seen[lower] > 1:
                text = f"{text}_{seen[lower]}"
            normalized.append(text)
        return normalized

    def _read_rows(self, table_name: str) -> List[Dict[str, Any]]:
        if self._mode == "remote_dataset":
            entry = self._resolve_dataset_entry(table_name)
            storage_path = Path(str(entry.get("storage_key")))
            file_type = str(entry.get("file_type") or storage_path.suffix.lstrip(".")).lower()
            sheet_name = str(entry.get("sheet_name") or "").strip() or None
            _, rows = self._read_rows_from_file(storage_path, file_type, sheet_name)
            return rows

        header_row = self._header_row()
        sheet_name = self._resolve_sheet_name(table_name)
        if self._file_type == "xlsx":
            sheet = self._workbook[sheet_name]
            raw_rows = list(sheet.iter_rows(values_only=True))
            if not raw_rows:
                return []
            header_idx = min(header_row - 1, max(len(raw_rows) - 1, 0))
            headers = self._normalize_headers(list(raw_rows[header_idx]))
            data_rows = raw_rows[header_idx + 1 :]
        else:
            sheet = self._workbook.sheet_by_name(sheet_name)
            if sheet.nrows <= 0:
                return []
            header_idx = min(header_row - 1, max(sheet.nrows - 1, 0))
            headers = self._normalize_headers(sheet.row_values(header_idx))
            data_rows = [sheet.row_values(i) for i in range(header_idx + 1, sheet.nrows)]

        records: List[Dict[str, Any]] = []
        for row in data_rows:
            if len(row) < len(headers):
                row = list(row) + [None] * (len(headers) - len(row))
            values = row[: len(headers)]
            records.append(dict(zip(headers, values)))
        return records

    def _infer_data_type(self, values: List[Any]) -> str:
        for value in values:
            if value is None:
                continue
            if isinstance(value, bool):
                return "BOOLEAN"
            if isinstance(value, int):
                return "INTEGER"
            if isinstance(value, float):
                return "FLOAT"
            if isinstance(value, (datetime, date)):
                return "DATETIME"
            return "STRING"
        return "STRING"

    def get_columns(self, table_name: str) -> List[ColumnInfo]:
        if self._mode == "remote_dataset":
            entry = self._resolve_dataset_entry(table_name)
            storage_path = Path(str(entry.get("storage_key")))
            file_type = str(entry.get("file_type") or storage_path.suffix.lstrip(".")).lower()
            sheet_name = str(entry.get("sheet_name") or "").strip() or None
            headers, rows = self._read_rows_from_file(storage_path, file_type, sheet_name)
            columns: List[ColumnInfo] = []
            for header in headers:
                sample_values = [row.get(header) for row in rows[:100]]
                columns.append(
                    ColumnInfo(
                        name=header,
                        data_type=self._infer_data_type(sample_values),
                        nullable=True,
                        is_primary_key=False,
                    )
                )
            return columns

        rows = self._read_rows(table_name)
        headers = list(rows[0].keys()) if rows else self._normalize_headers([])
        if not headers:
            sheet_name = self._resolve_sheet_name(table_name)
            if self._file_type == "xlsx":
                sheet = self._workbook[sheet_name]
                raw_rows = list(sheet.iter_rows(values_only=True))
                if raw_rows:
                    headers = self._normalize_headers(list(raw_rows[min(self._header_row() - 1, len(raw_rows) - 1)]))
            else:
                sheet = self._workbook.sheet_by_name(sheet_name)
                if sheet.nrows > 0:
                    headers = self._normalize_headers(sheet.row_values(min(self._header_row() - 1, sheet.nrows - 1)))

        columns: List[ColumnInfo] = []
        for header in headers:
            sample_values = [row.get(header) for row in rows[:100]]
            columns.append(
                ColumnInfo(
                    name=header,
                    data_type=self._infer_data_type(sample_values),
                    nullable=True,
                    is_primary_key=False,
                )
            )
        return columns

    def get_indexes(self, table_name: str) -> List[IndexInfo]:
        return []

    def get_constraints(self, table_name: str) -> List[ConstraintInfo]:
        return []

    def get_primary_keys(self, table_name: str) -> List[str]:
        return []

    def get_row_count(self, table_name: str, where_clause: str = None) -> int:
        if where_clause:
            raise ValueError("Excel 数据源不支持 where 过滤")
        return len(self._read_rows(table_name))

    def fetch_data(
        self,
        table_name: str,
        columns: List[str] = None,
        where_clause: str = None,
        order_by: List[str] = None,
        offset: int = 0,
        limit: int = 1000,
    ) -> List[Dict[str, Any]]:
        if where_clause:
            raise ValueError("Excel 数据源不支持 where 过滤")

        rows = self._read_rows(table_name)
        if order_by:
            for key in reversed(order_by):
                field = str(key).split()[0]
                rows = sorted(rows, key=lambda item: (item.get(field) is None, str(item.get(field))))

        sliced = rows[offset : offset + limit]
        if not columns:
            return sliced
        return [{c: row.get(c) for c in columns} for row in sliced]

    def get_version(self) -> str:
        if self._mode == "remote_dataset":
            return "Excel (dataset)"
        file_type = self._file_type or (self._storage_path.suffix.lstrip(".") if self._storage_path else "unknown")
        return f"Excel ({file_type})"
