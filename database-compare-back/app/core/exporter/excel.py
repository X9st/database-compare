"""Excel导出器"""
from typing import List, Dict, Any
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel报告导出器"""
    
    def __init__(self):
        self.wb = Workbook()
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_compare_result(self, result: Dict[str, Any], structure_diffs: List[Dict],
                              data_diffs: List[Dict], output_path: str) -> str:
        """导出比对结果为Excel文件"""
        # 删除默认sheet
        del self.wb['Sheet']
        
        # 创建汇总sheet
        self._create_summary_sheet(result)
        
        # 创建结构差异sheet
        if structure_diffs:
            self._create_structure_diffs_sheet(structure_diffs)
        
        # 创建数据差异sheet
        if data_diffs:
            self._create_data_diffs_sheet(data_diffs)
        
        # 保存文件
        self.wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, result: Dict[str, Any]):
        """创建汇总sheet"""
        ws = self.wb.create_sheet("汇总", 0)
        
        # 标题
        ws.merge_cells('A1:D1')
        ws['A1'] = "数据库比对报告"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        info_data = [
            ["报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["源数据库", result.get('source_db', {}).get('name', '')],
            ["目标数据库", result.get('target_db', {}).get('name', '')],
            ["比对耗时", f"{result.get('duration_seconds', 0)} 秒"],
        ]
        
        row = 3
        for info in info_data:
            ws.cell(row=row, column=1, value=info[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=info[1])
            row += 1
        
        # 汇总统计
        row += 1
        ws.cell(row=row, column=1, value="汇总统计").font = Font(bold=True, size=12)
        row += 1
        
        summary = result.get('summary', {})
        summary_data = [
            ["比对表数量", summary.get('total_tables', 0)],
            ["结构一致表", summary.get('structure_match_tables', 0)],
            ["结构差异表", summary.get('structure_diff_tables', 0)],
            ["数据一致表", summary.get('data_match_tables', 0)],
            ["数据差异表", summary.get('data_diff_tables', 0)],
            ["结构差异数", summary.get('total_structure_diffs', 0)],
            ["数据差异数", summary.get('total_data_diffs', 0)],
        ]
        
        for data in summary_data:
            ws.cell(row=row, column=1, value=data[0])
            ws.cell(row=row, column=2, value=data[1])
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_structure_diffs_sheet(self, diffs: List[Dict]):
        """创建结构差异sheet"""
        ws = self.wb.create_sheet("结构差异")
        
        # 表头
        headers = ["表名", "差异类型", "字段名", "源值", "目标值", "差异描述"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        for row, diff in enumerate(diffs, 2):
            ws.cell(row=row, column=1, value=diff.get('table_name', '')).border = self.thin_border
            ws.cell(row=row, column=2, value=diff.get('diff_type', '')).border = self.thin_border
            ws.cell(row=row, column=3, value=diff.get('field_name', '')).border = self.thin_border
            ws.cell(row=row, column=4, value=diff.get('source_value', '')).border = self.thin_border
            ws.cell(row=row, column=5, value=diff.get('target_value', '')).border = self.thin_border
            ws.cell(row=row, column=6, value=diff.get('diff_detail', '')).border = self.thin_border
        
        # 调整列宽
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_data_diffs_sheet(self, diffs: List[Dict]):
        """创建数据差异sheet"""
        ws = self.wb.create_sheet("数据差异")
        
        # 表头
        headers = ["表名", "主键", "差异类型", "差异字段", "源值", "目标值"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # 数据
        for row, diff in enumerate(diffs, 2):
            ws.cell(row=row, column=1, value=diff.get('table_name', '')).border = self.thin_border
            ws.cell(row=row, column=2, value=str(diff.get('primary_key', {}))).border = self.thin_border
            ws.cell(row=row, column=3, value=diff.get('diff_type', '')).border = self.thin_border
            ws.cell(row=row, column=4, value=str(diff.get('diff_columns', []))).border = self.thin_border
            ws.cell(row=row, column=5, value=str(diff.get('source_values', {}))).border = self.thin_border
            ws.cell(row=row, column=6, value=str(diff.get('target_values', {}))).border = self.thin_border
        
        # 调整列宽
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 25
