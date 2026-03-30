"""数据源服务"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any, Dict, List, Optional, Tuple
import shutil
import stat
import uuid

from sqlalchemy.orm import Session

from app.models.datasource import DataSource, DataSourceGroup
from app.schemas.datasource import (
    CreateDataSourceRequest,
    CreateGroupRequest,
    CreateRemoteDatasetRequest,
    DataSourceGroupResponse,
    DataSourceResponse,
    FileUploadResponse,
    RemoteDatasetRefreshResponse,
    TableInfo,
    TableSchema,
    TestConnectionRequest,
    TestConnectionResult,
    UpdateDataSourceRequest,
    UpdateGroupRequest,
    ColumnInfo,
    IndexInfo,
    ConstraintInfo,
)
from app.core.connector import ConnectorFactory
from app.utils.crypto import encrypt, decrypt


class DataSourceService:
    """数据源服务"""

    FILE_DB_TYPES = {"excel", "dbf"}
    FILE_MODES = {"single_file", "remote_dataset"}
    ALLOWED_UPLOAD_TYPES = {
        ".xlsx": "xlsx",
        ".xls": "xls",
        ".dbf": "dbf",
    }
    DEFAULT_FILE_CONNECTION = {
        "host": "local-file",
        "port": 0,
        "database": "local_file",
        "username": "file_user",
        "password": "",
    }

    def __init__(self, db: Session):
        self.db = db
        self.upload_dir = Path("data/uploads/datasources")
        self.dataset_root_dir = Path("data/uploads/datasets")

    def upload_datasource_file(self, filename: str, content: bytes) -> FileUploadResponse:
        """保存数据源文件并返回元信息"""
        safe_name = Path(filename or "").name
        if not safe_name:
            raise ValueError("文件名不能为空")

        suffix = Path(safe_name).suffix.lower()
        if suffix not in self.ALLOWED_UPLOAD_TYPES:
            raise ValueError("仅支持上传 .xlsx/.xls/.dbf 文件")

        self.upload_dir.mkdir(parents=True, exist_ok=True)
        file_id = uuid.uuid4().hex
        storage_name = f"{file_id}{suffix}"
        storage_path = self.upload_dir / storage_name
        storage_path.write_bytes(content)

        return FileUploadResponse(
            file_id=file_id,
            storage_key=str(storage_path.as_posix()),
            original_name=safe_name,
            file_type=self.ALLOWED_UPLOAD_TYPES[suffix],
            file_size=len(content),
        )

    def _resolve_file_mode(self, extra_config: Dict[str, Any]) -> str:
        mode = str((extra_config or {}).get("mode") or "").strip().lower()
        if mode:
            return mode
        return "remote_dataset" if isinstance((extra_config or {}).get("sftp"), dict) else "single_file"

    def _encrypt_remote_dataset_password(self, extra_config: Dict[str, Any]) -> Dict[str, Any]:
        result = dict(extra_config or {})
        sftp = dict(result.get("sftp") or {})
        plain_password = str(sftp.get("password") or "").strip()
        if plain_password:
            sftp["password_encrypted"] = encrypt(plain_password)
        sftp.pop("password", None)
        sftp["transport"] = "sftp"
        result["sftp"] = sftp
        result["mode"] = "remote_dataset"
        return result

    def _db_payload_from_input(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        db_type = payload["db_type"]
        if db_type in self.FILE_DB_TYPES:
            file_cfg = dict(payload.get("extra_config") or {})
            mode = self._resolve_file_mode(file_cfg)
            file_cfg["mode"] = mode
            if mode == "remote_dataset":
                file_cfg = self._encrypt_remote_dataset_password(file_cfg)

            original_name = str(file_cfg.get("original_name") or payload.get("name") or "file_source")
            return {
                "host": self.DEFAULT_FILE_CONNECTION["host"],
                "port": self.DEFAULT_FILE_CONNECTION["port"],
                "database": payload.get("database") or original_name,
                "schema": payload.get("schema"),
                "username": self.DEFAULT_FILE_CONNECTION["username"],
                "password": payload.get("password") or self.DEFAULT_FILE_CONNECTION["password"],
                "extra_config": file_cfg,
            }
        return {
            "host": payload["host"],
            "port": payload["port"],
            "database": payload["database"],
            "schema": payload.get("schema"),
            "username": payload["username"],
            "password": payload.get("password"),
            "extra_config": payload.get("extra_config"),
        }

    def _connector_kwargs_from_model(self, ds: DataSource, schema: Optional[str] = None) -> Dict[str, Any]:
        password = ""
        if ds.password_encrypted:
            try:
                password = decrypt(ds.password_encrypted)
            except Exception:
                password = ""
        return {
            "db_type": ds.db_type,
            "host": ds.host,
            "port": ds.port,
            "database": ds.database,
            "schema": schema if schema is not None else ds.schema,
            "username": ds.username,
            "password": password,
            "charset": ds.charset,
            "timeout": ds.timeout,
            "extra_config": ds.extra_config or {},
        }

    def _mask_extra_config_for_response(self, extra_config: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        if not isinstance(extra_config, dict):
            return None
        result = dict(extra_config)
        if self._resolve_file_mode(result) == "remote_dataset":
            sftp = dict(result.get("sftp") or {})
            sftp.pop("password_encrypted", None)
            result["sftp"] = sftp
        return result

    def _check_remote_dataset_mode(self, ds: DataSource) -> Dict[str, Any]:
        if (ds.db_type or "").lower() not in self.FILE_DB_TYPES:
            raise ValueError("仅文件型数据源支持刷新数据集")
        extra_config = dict(ds.extra_config or {})
        if self._resolve_file_mode(extra_config) != "remote_dataset":
            raise ValueError("当前数据源不是远程目录数据集")
        return extra_config

    def _normalize_table_name(self, raw_name: str, used: Dict[str, int]) -> str:
        base = str(raw_name or "").strip() or "table"
        key = base.lower()
        count = used.get(key, 0)
        used[key] = count + 1
        if count == 0:
            return base
        return f"{base}_{count + 1}"

    def _scan_excel_tables(
        self,
        file_path: Path,
        file_name: str,
        table_index: Dict[str, Dict[str, Any]],
        used_names: Dict[str, int],
    ) -> int:
        suffix = file_path.suffix.lower()
        stem = file_path.stem
        sheet_names: List[str]
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, data_only=True, read_only=True)
            try:
                sheet_names = list(workbook.sheetnames)
            finally:
                workbook.close()
        elif suffix == ".xls":
            import xlrd

            workbook = xlrd.open_workbook(file_path)
            sheet_names = list(workbook.sheet_names())
        else:
            raise ValueError(f"不支持的 Excel 文件类型: {suffix}")

        for sheet_name in sheet_names:
            table_name = self._normalize_table_name(f"{stem}__{sheet_name}", used_names)
            table_index[table_name] = {
                "storage_key": str(file_path.resolve().as_posix()),
                "original_name": file_name,
                "file_type": suffix.lstrip("."),
                "sheet_name": sheet_name,
            }
        return len(sheet_names)

    def _scan_dbf_table(
        self,
        file_path: Path,
        file_name: str,
        table_index: Dict[str, Dict[str, Any]],
        used_names: Dict[str, int],
    ) -> int:
        table_name = self._normalize_table_name(file_path.stem, used_names)
        table_index[table_name] = {
            "storage_key": str(file_path.resolve().as_posix()),
            "original_name": file_name,
            "file_type": "dbf",
        }
        return 1

    def _list_remote_source_files(self, sftp, base_dir: str, db_type: str) -> List[str]:
        try:
            entries = sftp.listdir_attr(base_dir)
        except Exception as exc:
            raise ValueError(f"读取远程目录失败: {exc}") from exc

        matched: List[str] = []
        for entry in entries:
            if not stat.S_ISREG(getattr(entry, "st_mode", 0)):
                continue
            file_name = Path(str(getattr(entry, "filename", ""))).name
            suffix = Path(file_name).suffix.lower()
            if db_type == "excel" and suffix in {".xlsx", ".xls"}:
                matched.append(file_name)
            if db_type == "dbf" and suffix == ".dbf":
                matched.append(file_name)
        return sorted(matched, key=str.lower)

    def _remote_sftp_connection_test(self, db_type: str, extra_config: Dict[str, Any]) -> TestConnectionResult:
        sftp_cfg = dict((extra_config or {}).get("sftp") or {})
        host = str(sftp_cfg.get("host") or "").strip()
        username = str(sftp_cfg.get("username") or "").strip()
        base_dir = str(sftp_cfg.get("base_dir") or "").strip()
        port = int(sftp_cfg.get("port") or 22)
        password = str(sftp_cfg.get("password") or "").strip()
        if not password:
            encrypted = str(sftp_cfg.get("password_encrypted") or "").strip()
            if encrypted:
                try:
                    password = decrypt(encrypted)
                except Exception:
                    password = ""
        if not host or not username or not base_dir or not password:
            return TestConnectionResult(success=False, message="远程数据集 SFTP 配置不完整")

        transport = None
        sftp = None
        try:
            import paramiko

            transport = paramiko.Transport((host, port))
            transport.connect(username=username, password=password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            matched = self._list_remote_source_files(sftp=sftp, base_dir=base_dir, db_type=db_type)
            return TestConnectionResult(
                success=True,
                message=f"连接成功，目录中匹配文件 {len(matched)} 个",
                version="SFTP",
            )
        except Exception as exc:
            return TestConnectionResult(success=False, message=f"SFTP 连接失败: {exc}")
        finally:
            if sftp is not None:
                try:
                    sftp.close()
                except Exception:
                    pass
            if transport is not None:
                try:
                    transport.close()
                except Exception:
                    pass

    def _build_remote_snapshot(
        self,
        ds: DataSource,
        extra_config: Dict[str, Any],
        plain_password: Optional[str] = None,
    ) -> Tuple[Dict[str, Any], Optional[Path]]:
        sftp_cfg = dict(extra_config.get("sftp") or {})
        host = str(sftp_cfg.get("host") or "").strip()
        username = str(sftp_cfg.get("username") or "").strip()
        base_dir = str(sftp_cfg.get("base_dir") or "").strip()
        port = int(sftp_cfg.get("port") or 22)
        password = str(plain_password or "").strip()
        if not password:
            encrypted = str(sftp_cfg.get("password_encrypted") or "").strip()
            if encrypted:
                try:
                    password = decrypt(encrypted)
                except Exception:
                    password = ""

        if not host or not username or not base_dir or not password:
            raise ValueError("SFTP 连接信息不完整，无法刷新远程数据集")

        dataset_dir = (self.dataset_root_dir / ds.id).resolve()
        dataset_dir.mkdir(parents=True, exist_ok=True)
        snapshot_dir = dataset_dir / f"snapshot_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
        snapshot_dir.mkdir(parents=True, exist_ok=True)

        table_index: Dict[str, Dict[str, Any]] = {}
        used_names: Dict[str, int] = {}
        failed_files: List[Dict[str, Any]] = []
        success_files = 0
        table_count = 0

        transport = None
        sftp = None
        try:
            import paramiko

            transport = paramiko.Transport((host, port))
            transport.connect(username=username, password=password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            source_files = self._list_remote_source_files(sftp=sftp, base_dir=base_dir, db_type=ds.db_type)
            for file_name in source_files:
                remote_path = str(PurePosixPath(base_dir) / file_name)
                local_path = snapshot_dir / file_name
                try:
                    sftp.get(remote_path, str(local_path))
                    if ds.db_type == "excel":
                        table_count += self._scan_excel_tables(
                            file_path=local_path,
                            file_name=file_name,
                            table_index=table_index,
                            used_names=used_names,
                        )
                    else:
                        table_count += self._scan_dbf_table(
                            file_path=local_path,
                            file_name=file_name,
                            table_index=table_index,
                            used_names=used_names,
                        )
                    success_files += 1
                except Exception as exc:
                    failed_files.append({"file_name": file_name, "error": str(exc)})
                    try:
                        local_path.unlink(missing_ok=True)
                    except Exception:
                        pass
        except Exception as exc:
            raise ValueError(f"SFTP 刷新失败: {exc}") from exc
        finally:
            if sftp is not None:
                try:
                    sftp.close()
                except Exception:
                    pass
            if transport is not None:
                try:
                    transport.close()
                except Exception:
                    pass

        if success_files <= 0 or table_count <= 0:
            shutil.rmtree(snapshot_dir, ignore_errors=True)
            raise ValueError("本次刷新没有成功导入任何文件，已保留上一版快照")

        snapshot = {
            "dataset_root": str(snapshot_dir.as_posix()),
            "table_index": table_index,
            "file_count": success_files,
            "table_count": table_count,
            "failed_files": failed_files,
            "last_refresh_at": datetime.utcnow().isoformat(),
        }

        old_snapshot_dir = None
        old_snapshot = (extra_config.get("snapshot") or {})
        old_root = str(old_snapshot.get("dataset_root") or "").strip()
        if old_root:
            old_snapshot_dir = Path(old_root)
            if not old_snapshot_dir.is_absolute():
                old_snapshot_dir = Path.cwd() / old_snapshot_dir
            old_snapshot_dir = old_snapshot_dir.resolve()
        return snapshot, old_snapshot_dir

    def create_remote_dataset(self, data: CreateRemoteDatasetRequest) -> DataSourceResponse:
        payload = data.model_dump(exclude_unset=True)
        conn_payload = self._db_payload_from_input(payload)
        ds = DataSource(
            id=str(uuid.uuid4()),
            name=payload["name"],
            group_id=payload.get("group_id"),
            db_type=payload["db_type"],
            host=conn_payload["host"],
            port=conn_payload["port"],
            database=conn_payload["database"],
            schema=None,
            username=conn_payload["username"],
            password_encrypted=encrypt(""),
            charset=payload.get("charset", "UTF-8"),
            timeout=payload.get("timeout", 30),
            extra_config=conn_payload.get("extra_config"),
        )
        self.db.add(ds)
        self.db.commit()
        self.db.refresh(ds)

        plain_password = str((((payload.get("extra_config") or {}).get("sftp") or {}).get("password")) or "")
        try:
            self.refresh_remote_dataset(ds_id=ds.id, plain_password=plain_password)
        except Exception as exc:
            self.db.delete(ds)
            self.db.commit()
            raise ValueError(f"远程数据集首次导入失败: {exc}") from exc
        return self.get_by_id(ds.id)

    def refresh_remote_dataset(
        self,
        ds_id: str,
        plain_password: Optional[str] = None,
    ) -> RemoteDatasetRefreshResponse:
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            raise ValueError("数据源不存在")

        extra_config = self._check_remote_dataset_mode(ds)
        snapshot, old_snapshot_dir = self._build_remote_snapshot(
            ds=ds,
            extra_config=extra_config,
            plain_password=plain_password,
        )

        extra_config["mode"] = "remote_dataset"
        extra_config["snapshot"] = snapshot
        ds.extra_config = extra_config
        self.db.commit()
        self.db.refresh(ds)

        if old_snapshot_dir and old_snapshot_dir.exists():
            safe_parent = (self.dataset_root_dir / ds.id).resolve()
            current_snapshot_dir = Path(str(snapshot["dataset_root"])).resolve()
            try:
                if safe_parent in old_snapshot_dir.parents and old_snapshot_dir != current_snapshot_dir:
                    shutil.rmtree(old_snapshot_dir, ignore_errors=True)
            except Exception:
                pass

        return RemoteDatasetRefreshResponse(
            datasource_id=ds.id,
            file_count=int(snapshot.get("file_count") or 0),
            table_count=int(snapshot.get("table_count") or 0),
            failed_files=list(snapshot.get("failed_files") or []),
            last_refresh_at=datetime.fromisoformat(str(snapshot.get("last_refresh_at"))),
        )
    
    # ==================== 数据源操作 ====================
    
    def get_list(self, group_id: str = None, keyword: str = None, 
                 db_type: str = None) -> List[DataSourceResponse]:
        """获取数据源列表"""
        query = self.db.query(DataSource)
        
        if group_id:
            query = query.filter(DataSource.group_id == group_id)
        if db_type:
            query = query.filter(DataSource.db_type == db_type)
        if keyword:
            query = query.filter(
                DataSource.name.contains(keyword) | 
                DataSource.host.contains(keyword) |
                DataSource.database.contains(keyword)
            )
        
        datasources = query.order_by(DataSource.created_at.desc()).all()
        
        result = []
        for ds in datasources:
            group_name = None
            if ds.group_id:
                group = self.db.query(DataSourceGroup).filter(
                    DataSourceGroup.id == ds.group_id
                ).first()
                group_name = group.name if group else None
            
            result.append(DataSourceResponse(
                id=ds.id,
                name=ds.name,
                group_id=ds.group_id,
                group_name=group_name,
                db_type=ds.db_type,
                host=ds.host,
                port=ds.port,
                database=ds.database,
                schema=ds.schema,
                username=ds.username,
                charset=ds.charset,
                timeout=ds.timeout,
                extra_config=self._mask_extra_config_for_response(ds.extra_config),
                created_at=ds.created_at,
                updated_at=ds.updated_at
            ))
        
        return result
    
    def get_by_id(self, ds_id: str) -> Optional[DataSourceResponse]:
        """获取单个数据源"""
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            return None
        
        group_name = None
        if ds.group_id:
            group = self.db.query(DataSourceGroup).filter(
                DataSourceGroup.id == ds.group_id
            ).first()
            group_name = group.name if group else None
        
        return DataSourceResponse(
            id=ds.id,
            name=ds.name,
            group_id=ds.group_id,
            group_name=group_name,
            db_type=ds.db_type,
            host=ds.host,
            port=ds.port,
            database=ds.database,
            schema=ds.schema,
            username=ds.username,
            charset=ds.charset,
            timeout=ds.timeout,
            extra_config=self._mask_extra_config_for_response(ds.extra_config),
            created_at=ds.created_at,
            updated_at=ds.updated_at
        )
    
    def create(self, data: CreateDataSourceRequest) -> DataSourceResponse:
        """创建数据源"""
        payload = data.model_dump(exclude_unset=True)
        conn_payload = self._db_payload_from_input(payload)
        ds = DataSource(
            id=str(uuid.uuid4()),
            name=payload["name"],
            group_id=payload.get("group_id"),
            db_type=payload["db_type"],
            host=conn_payload["host"],
            port=conn_payload["port"],
            database=conn_payload["database"],
            schema=conn_payload.get("schema"),
            username=conn_payload["username"],
            password_encrypted=encrypt(conn_payload.get("password") or ""),
            charset=payload.get("charset", "UTF-8"),
            timeout=payload.get("timeout", 30),
            extra_config=conn_payload.get("extra_config"),
        )
        
        self.db.add(ds)
        self.db.commit()
        self.db.refresh(ds)
        
        return self.get_by_id(ds.id)
    
    def update(self, ds_id: str, data: UpdateDataSourceRequest) -> Optional[DataSourceResponse]:
        """更新数据源"""
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            return None
        
        update_data = data.model_dump(exclude_unset=True)

        # 处理密码加密
        if 'password' in update_data:
            update_data['password_encrypted'] = encrypt(update_data.pop('password'))

        if "extra_config" in update_data and isinstance(update_data["extra_config"], dict):
            incoming_extra = dict(update_data["extra_config"])
            merged_extra = dict(ds.extra_config or {})
            mode = self._resolve_file_mode(incoming_extra)
            incoming_extra["mode"] = mode
            if mode == "remote_dataset":
                incoming_sftp = dict(incoming_extra.get("sftp") or {})
                merged_sftp = dict(merged_extra.get("sftp") or {})
                if not incoming_sftp.get("password") and not incoming_sftp.get("password_encrypted"):
                    incoming_sftp["password_encrypted"] = merged_sftp.get("password_encrypted")
                incoming_extra["sftp"] = {**merged_sftp, **incoming_sftp}
                incoming_extra = self._encrypt_remote_dataset_password(incoming_extra)
            merged_extra.update(incoming_extra)
            if mode == "remote_dataset":
                merged_sftp = dict(merged_extra.get("sftp") or {})
                if not str(merged_sftp.get("password_encrypted") or "").strip():
                    raise ValueError("远程目录数据集更新必须提供 sftp.password")
            update_data["extra_config"] = merged_extra

        db_type = update_data.get("db_type", ds.db_type)
        if db_type in self.FILE_DB_TYPES:
            if "host" not in update_data:
                update_data["host"] = self.DEFAULT_FILE_CONNECTION["host"]
            if "port" not in update_data:
                update_data["port"] = self.DEFAULT_FILE_CONNECTION["port"]
            if "username" not in update_data:
                update_data["username"] = self.DEFAULT_FILE_CONNECTION["username"]
            if "database" not in update_data:
                update_data["database"] = ds.database or self.DEFAULT_FILE_CONNECTION["database"]
            if "password_encrypted" not in update_data:
                update_data["password_encrypted"] = encrypt("")
        
        for key, value in update_data.items():
            if hasattr(ds, key):
                setattr(ds, key, value)
        
        self.db.commit()
        self.db.refresh(ds)
        
        return self.get_by_id(ds.id)
    
    def delete(self, ds_id: str) -> bool:
        """删除数据源"""
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            return False

        extra_config = dict(ds.extra_config or {})
        self.db.delete(ds)
        self.db.commit()

        if self._resolve_file_mode(extra_config) == "remote_dataset":
            dataset_dir = (self.dataset_root_dir / ds_id).resolve()
            shutil.rmtree(dataset_dir, ignore_errors=True)
        return True
    
    def test_connection_by_id(self, ds_id: str) -> TestConnectionResult:
        """测试已保存数据源的连接"""
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            return TestConnectionResult(
                success=False,
                message="数据源不存在"
            )

        extra_config = dict(ds.extra_config or {})
        if (ds.db_type or "").lower() in self.FILE_DB_TYPES and self._resolve_file_mode(extra_config) == "remote_dataset":
            return self._remote_sftp_connection_test(db_type=ds.db_type, extra_config=extra_config)

        return self._test_connection(
            **self._connector_kwargs_from_model(ds)
        )
    
    def test_connection_direct(self, data: TestConnectionRequest) -> TestConnectionResult:
        """直接测试连接（不保存）"""
        payload = data.model_dump(exclude_unset=True)
        db_payload = self._db_payload_from_input(payload)
        extra_config = db_payload.get("extra_config") or {}
        if payload["db_type"] in self.FILE_DB_TYPES and self._resolve_file_mode(extra_config) == "remote_dataset":
            return self._remote_sftp_connection_test(
                db_type=payload["db_type"],
                extra_config=extra_config,
            )
        return self._test_connection(
            db_type=payload["db_type"],
            host=db_payload["host"],
            port=db_payload["port"],
            database=db_payload["database"],
            schema=db_payload.get("schema"),
            username=db_payload["username"],
            password=db_payload.get("password") or "",
            charset=payload.get("charset", "UTF-8"),
            timeout=payload.get("timeout", 30),
            extra_config=extra_config,
        )
    
    def _test_connection(self, **kwargs) -> TestConnectionResult:
        """测试连接实现"""
        try:
            connector = ConnectorFactory.create(**kwargs)
            result = connector.test_connection()
            return TestConnectionResult(**result)
        except ValueError as e:
            return TestConnectionResult(
                success=False,
                message=str(e)
            )
        except Exception as e:
            return TestConnectionResult(
                success=False,
                message=f"连接失败: {str(e)}"
            )
    
    def get_tables(self, ds_id: str, schema: str = None, keyword: str = None) -> List[TableInfo]:
        """获取表列表"""
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            return []
        
        try:
            connector = ConnectorFactory.create(
                **self._connector_kwargs_from_model(ds, schema=schema or ds.schema)
            )
            
            with connector:
                tables = connector.get_tables()
                if keyword:
                    keyword_lower = keyword.lower()
                    tables = [t for t in tables if keyword_lower in t.name.lower()]
                return [TableInfo(
                    name=t.name,
                    schema=t.schema,
                    comment=t.comment,
                    row_count=t.row_count
                ) for t in tables]
        except Exception as e:
            raise ValueError(f"获取表列表失败: {str(e)}")
    
    def get_table_schema(self, ds_id: str, table_name: str) -> Optional[TableSchema]:
        """获取表结构"""
        ds = self.db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds:
            return None
        
        try:
            connector = ConnectorFactory.create(
                **self._connector_kwargs_from_model(ds)
            )
            
            with connector:
                columns = connector.get_columns(table_name)
                indexes = connector.get_indexes(table_name)
                constraints = connector.get_constraints(table_name)
                table_comment = None
                try:
                    table_meta = connector.get_tables()
                    match = next(
                        (t for t in table_meta if str(t.name).lower() == str(table_name).lower()),
                        None,
                    )
                    table_comment = match.comment if match else None
                except Exception:
                    table_comment = None
                
                return TableSchema(
                    table_name=table_name,
                    comment=table_comment,
                    columns=[ColumnInfo(
                        name=c.name,
                        data_type=c.data_type,
                        length=c.length,
                        precision=c.precision,
                        scale=c.scale,
                        nullable=c.nullable,
                        default_value=c.default_value,
                        comment=c.comment,
                        is_primary_key=c.is_primary_key
                    ) for c in columns],
                    indexes=[IndexInfo(
                        name=i.name,
                        columns=i.columns,
                        is_unique=i.is_unique,
                        is_primary=i.is_primary,
                        index_type=i.index_type
                    ) for i in indexes],
                    constraints=[ConstraintInfo(
                        name=c.name,
                        constraint_type=c.constraint_type,
                        columns=c.columns,
                        reference_table=c.reference_table,
                        reference_columns=c.reference_columns
                    ) for c in constraints]
                )
        except Exception as e:
            raise ValueError(f"获取表结构失败: {str(e)}")
    
    def get_datasource_model(self, ds_id: str) -> Optional[DataSource]:
        """获取数据源模型（内部使用）"""
        return self.db.query(DataSource).filter(DataSource.id == ds_id).first()
    
    # ==================== 分组操作 ====================
    
    def get_groups(self) -> List[DataSourceGroupResponse]:
        """获取分组列表"""
        groups = self.db.query(DataSourceGroup).order_by(
            DataSourceGroup.sort_order, DataSourceGroup.name
        ).all()
        
        result = []
        for group in groups:
            count = self.db.query(DataSource).filter(
                DataSource.group_id == group.id
            ).count()
            
            result.append(DataSourceGroupResponse(
                id=group.id,
                name=group.name,
                count=count,
                sort_order=group.sort_order
            ))
        
        return result
    
    def create_group(self, data: CreateGroupRequest) -> DataSourceGroupResponse:
        """创建分组"""
        # 获取最大排序号
        max_order = self.db.query(DataSourceGroup).count()
        
        group = DataSourceGroup(
            id=str(uuid.uuid4()),
            name=data.name,
            sort_order=max_order + 1
        )
        
        self.db.add(group)
        self.db.commit()
        self.db.refresh(group)
        
        return DataSourceGroupResponse(
            id=group.id,
            name=group.name,
            count=0,
            sort_order=group.sort_order
        )
    
    def update_group(self, group_id: str, data: UpdateGroupRequest) -> Optional[DataSourceGroupResponse]:
        """更新分组"""
        group = self.db.query(DataSourceGroup).filter(
            DataSourceGroup.id == group_id
        ).first()
        
        if not group:
            return None
        
        if data.name is not None:
            group.name = data.name
        if data.sort_order is not None:
            group.sort_order = data.sort_order
        
        self.db.commit()
        self.db.refresh(group)
        
        count = self.db.query(DataSource).filter(
            DataSource.group_id == group.id
        ).count()
        
        return DataSourceGroupResponse(
            id=group.id,
            name=group.name,
            count=count,
            sort_order=group.sort_order
        )
    
    def delete_group(self, group_id: str) -> bool:
        """删除分组"""
        group = self.db.query(DataSourceGroup).filter(
            DataSourceGroup.id == group_id
        ).first()
        
        if not group:
            return False
        
        # 将该分组下的数据源的group_id置为null
        self.db.query(DataSource).filter(
            DataSource.group_id == group_id
        ).update({DataSource.group_id: None})
        
        self.db.delete(group)
        self.db.commit()
        return True
