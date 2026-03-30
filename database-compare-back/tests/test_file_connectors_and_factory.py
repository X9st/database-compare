"""文件连接器与工厂测试"""
from __future__ import annotations

import sys
import types
from pathlib import Path

import pytest

from app.core.comparator.structure import StructureComparator
from app.core.connector.base import ColumnInfo
from app.core.connector.excel import ExcelConnector
from app.core.connector.factory import ConnectorFactory
from app.core.connector.dbf import DBFConnector


def test_factory_supported_types_converged():
    assert set(ConnectorFactory.get_supported_types()) == {"mysql", "oracle", "dm", "inceptor", "excel", "dbf"}
    with pytest.raises(ValueError):
        ConnectorFactory.create(db_type="sqlserver", host="x", port=0, database="x", username="x", password="x")


def test_excel_connector_reads_xlsx(tmp_path: Path):
    from openpyxl import Workbook

    file_path = tmp_path / "orders.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(["id", "name"])
    ws.append([1, "alice"])
    ws.append([2, "bob"])
    wb.save(file_path)

    connector = ExcelConnector(
        host="local-file",
        port=0,
        database="orders",
        username="file_user",
        password="",
        extra_config={"storage_key": str(file_path), "file_type": "xlsx", "header_row": 1},
    )

    with connector:
        tables = connector.get_tables()
        assert [t.name for t in tables] == ["orders"]
        columns = connector.get_columns("orders")
        assert [c.name for c in columns] == ["id", "name"]
        rows = connector.fetch_data("orders", columns=["id", "name"], offset=1, limit=1)
        assert rows == [{"id": 2, "name": "bob"}]


def test_dbf_connector_with_stub_driver(monkeypatch, tmp_path: Path):
    file_path = tmp_path / "user.dbf"
    file_path.write_bytes(b"fake dbf")

    class FakeField:
        def __init__(self, name: str, typ: str, length: int = 0, decimal_count: int = 0):
            self.name = name
            self.type = typ
            self.length = length
            self.decimal_count = decimal_count

    class FakeDBF:
        def __init__(self, *_args, **_kwargs):
            self.fields = [FakeField("ID", "N", 18, 0), FakeField("NAME", "C", 50, 0)]
            self._rows = [{"ID": 1, "NAME": "Alice"}, {"ID": 2, "NAME": "Bob"}]

        def __iter__(self):
            return iter(self._rows)

    monkeypatch.setitem(sys.modules, "dbfread", types.SimpleNamespace(DBF=FakeDBF))

    connector = DBFConnector(
        host="local-file",
        port=0,
        database="user",
        username="file_user",
        password="",
        extra_config={"storage_key": str(file_path), "file_type": "dbf"},
    )
    with connector:
        tables = connector.get_tables()
        assert [t.name for t in tables] == ["user"]
        columns = connector.get_columns("user")
        assert [c.name for c in columns] == ["ID", "NAME"]
        rows = connector.fetch_data("user", columns=["ID"], offset=0, limit=1)
        assert rows == [{"ID": 1}]


def test_excel_connector_dataset_mode_reads_table_index(tmp_path: Path):
    from openpyxl import Workbook

    first = tmp_path / "sales_a.xlsx"
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Q1"
    ws1.append(["id", "amount"])
    ws1.append([1, 100])
    wb1.save(first)

    second = tmp_path / "sales_b.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Q2"
    ws2.append(["id", "amount"])
    ws2.append([2, 200])
    wb2.save(second)

    connector = ExcelConnector(
        host="local-file",
        port=0,
        database="sales",
        username="file_user",
        password="",
        extra_config={
            "mode": "remote_dataset",
            "snapshot": {
                "table_index": {
                    "sales_a__Q1": {
                        "storage_key": str(first),
                        "file_type": "xlsx",
                        "sheet_name": "Q1",
                        "original_name": "sales_a.xlsx",
                    },
                    "sales_b__Q2": {
                        "storage_key": str(second),
                        "file_type": "xlsx",
                        "sheet_name": "Q2",
                        "original_name": "sales_b.xlsx",
                    },
                }
            },
        },
    )

    with connector:
        tables = connector.get_tables()
        assert [t.name for t in tables] == ["sales_a__Q1", "sales_b__Q2"]
        columns = connector.get_columns("sales_b__Q2")
        assert [c.name for c in columns] == ["id", "amount"]
        rows = connector.fetch_data("sales_b__Q2")
        assert rows == [{"id": 2, "amount": 200}]


def test_dbf_connector_dataset_mode_with_stub_driver(monkeypatch, tmp_path: Path):
    first = tmp_path / "a.dbf"
    second = tmp_path / "b.dbf"
    first.write_bytes(b"fake")
    second.write_bytes(b"fake")

    class FakeField:
        def __init__(self, name: str, typ: str, length: int = 0, decimal_count: int = 0):
            self.name = name
            self.type = typ
            self.length = length
            self.decimal_count = decimal_count

    class FakeDBF:
        def __init__(self, path: str, *_args, **_kwargs):
            self.fields = [FakeField("ID", "N", 18, 0), FakeField("NAME", "C", 50, 0)]
            if path.endswith("a.dbf"):
                self._rows = [{"ID": 1, "NAME": "A"}]
            else:
                self._rows = [{"ID": 2, "NAME": "B"}]

        def __iter__(self):
            return iter(self._rows)

    monkeypatch.setitem(sys.modules, "dbfread", types.SimpleNamespace(DBF=FakeDBF))

    connector = DBFConnector(
        host="local-file",
        port=0,
        database="dbf-dataset",
        username="file_user",
        password="",
        extra_config={
            "mode": "remote_dataset",
            "snapshot": {
                "table_index": {
                    "a": {"storage_key": str(first), "file_type": "dbf", "original_name": "a.dbf"},
                    "b": {"storage_key": str(second), "file_type": "dbf", "original_name": "b.dbf"},
                }
            },
        },
    )

    with connector:
        tables = connector.get_tables()
        assert [t.name for t in tables] == ["a", "b"]
        rows = connector.fetch_data("b")
        assert rows == [{"ID": 2, "NAME": "B"}]


def test_structure_comparator_uses_lightweight_rules_for_file_source():
    class StubConn:
        def __init__(self, columns, is_file_source):
            self._columns = columns
            self.is_file_source = is_file_source

        def get_columns(self, _table):
            return self._columns

        def get_tables(self):
            return []

        def get_indexes(self, _table):
            return []

        def get_constraints(self, _table):
            return []

    source_conn = StubConn(
        [ColumnInfo(name="id", data_type="INTEGER"), ColumnInfo(name="name", data_type="STRING")],
        is_file_source=True,
    )
    target_conn = StubConn(
        [ColumnInfo(name="id", data_type="VARCHAR"), ColumnInfo(name="name", data_type="VARCHAR")],
        is_file_source=False,
    )
    comparator = StructureComparator(source_conn, target_conn, {"compare_comment": True, "compare_index": True, "compare_constraint": True})

    diffs = comparator.compare_table_structure("users")
    assert diffs == []
