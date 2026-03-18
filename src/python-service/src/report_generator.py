#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报告生成模块
支持 HTML 和 Excel 格式的比对报告导出
"""

import os
import json
from datetime import datetime
from typing import Dict, List
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportGenerator:
    """比对报告生成器"""
    
    def __init__(self, result: Dict, task_config: Dict):
        """
        初始化报告生成器
        
        Args:
            result: 比对结果数据
            task_config: 任务配置信息
        """
        self.result = result
        self.task_config = task_config
        self.generated_at = datetime.now()
    
    def generate_html(self) -> str:
        """生成 HTML 格式报告"""
        html_template = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>数据库比对报告</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #f5f5f5;
            padding: 20px;
            line-height: 1.6;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            margin-bottom: 8px;
            font-size: 28px;
        }
        .meta {
            color: #999;
            font-size: 14px;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 1px solid #eee;
        }
        .summary {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
            margin-bottom: 40px;
        }
        .summary-card {
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .summary-card.added {
            background: #f6ffed;
            border: 1px solid #b7eb8f;
        }
        .summary-card.removed {
            background: #fff2f0;
            border: 1px solid #ffccc7;
        }
        .summary-card.modified {
            background: #fffbe6;
            border: 1px solid #ffe58f;
        }
        .summary-card h3 {
            font-size: 14px;
            color: #666;
            margin-bottom: 12px;
        }
        .summary-card .number {
            font-size: 36px;
            font-weight: 600;
            color: #333;
        }
        .section {
            margin-bottom: 40px;
        }
        .section h2 {
            font-size: 20px;
            color: #333;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid #1890ff;
        }
        .diff-item {
            padding: 16px;
            margin-bottom: 12px;
            border-radius: 4px;
            border: 1px solid #e8e8e8;
        }
        .diff-item.added {
            background: #f6ffed;
            border-color: #b7eb8f;
        }
        .diff-item.removed {
            background: #fff2f0;
            border-color: #ffccc7;
        }
        .diff-item.modified {
            background: #fffbe6;
            border-color: #ffe58f;
        }
        .diff-header {
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 8px;
        }
        .badge {
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: 500;
            color: white;
        }
        .badge.added { background: #52c41a; }
        .badge.removed { background: #ff4d4f; }
        .badge.modified { background: #faad14; }
        .diff-name {
            font-weight: 600;
            font-size: 16px;
        }
        .diff-details {
            margin-left: 60px;
            font-size: 14px;
            color: #666;
        }
        .detail-row {
            margin-bottom: 4px;
        }
        .source { color: #ff4d4f; text-decoration: line-through; }
        .target { color: #52c41a; }
        .arrow { color: #999; margin: 0 8px; }
        .no-diff {
            text-align: center;
            color: #999;
            padding: 40px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>数据库比对报告</h1>
        <div class="meta">
            生成时间：{{ generated_at }} | 
            源数据库：{{ source_name }} | 
            目标数据库：{{ target_name }}
        </div>
        
        <div class="summary">
            <div class="summary-card added">
                <h3>新增</h3>
                <div class="number">{{ summary.added_tables }}</div>
                <div>表</div>
            </div>
            <div class="summary-card removed">
                <h3>删除</h3>
                <div class="number">{{ summary.removed_tables }}</div>
                <div>表</div>
            </div>
            <div class="summary-card modified">
                <h3>修改</h3>
                <div class="number">{{ summary.modified_tables }}</div>
                <div>表</div>
            </div>
        </div>
        
        <div class="section">
            <h2>表结构差异</h2>
            {% if table_diffs %}
                {% for diff in table_diffs %}
                <div class="diff-item {{ diff.type }}">
                    <div class="diff-header">
                        <span class="badge {{ diff.type }}">
                            {{ diff.type_label }}
                        </span>
                        <span class="diff-name">{{ diff.name }}</span>
                    </div>
                    {% if diff.differences %}
                    <div class="diff-details">
                        {% for d in diff.differences %}
                        <div class="detail-row">
                            {{ d.field }}: 
                            <span class="source">{{ d.source }}</span>
                            <span class="arrow">→</span>
                            <span class="target">{{ d.target }}</span>
                        </div>
                        {% endfor %}
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
            {% else %}
                <div class="no-diff">无表结构差异</div>
            {% endif %}
        </div>
        
        <div class="section">
            <h2>字段差异</h2>
            {% if column_diffs %}
                {% for table_name, diffs in column_diffs.items() %}
                <div class="diff-item modified">
                    <div class="diff-header">
                        <span class="badge modified">修改</span>
                        <span class="diff-name">{{ table_name }}</span>
                    </div>
                    <div class="diff-details">
                        {% if diffs.added %}
                            <div>新增字段: {{ diffs.added | join(', ') }}</div>
                        {% endif %}
                        {% if diffs.removed %}
                            <div>删除字段: {{ diffs.removed | join(', ') }}</div>
                        {% endif %}
                        {% if diffs.modified %}
                            <div>修改字段: {{ diffs.modified | join(', ') }}</div>
                        {% endif %}
                    </div>
                </div>
                {% endfor %}
            {% else %}
                <div class="no-diff">无字段差异</div>
            {% endif %}
        </div>
    </div>
</body>
</html>
        '''
        
        # 准备模板数据
        template_data = self._prepare_template_data()
        
        template = Template(html_template)
        return template.render(**template_data)
    
    def generate_excel(self, output_path: str):
        """生成 Excel 格式报告"""
        wb = Workbook()
        
        # 摘要页
        ws_summary = wb.active
        ws_summary.title = "摘要"
        self._create_summary_sheet(ws_summary)
        
        # 表结构差异页
        ws_tables = wb.create_sheet("表结构差异")
        self._create_table_diff_sheet(ws_tables)
        
        # 字段差异页
        ws_columns = wb.create_sheet("字段差异")
        self._create_column_diff_sheet(ws_columns)
        
        wb.save(output_path)
    
    def _prepare_template_data(self) -> Dict:
        """准备模板数据"""
        tables = self.result.get('tables', {})
        columns = self.result.get('columns', {})
        
        # 统计信息
        summary = {
            'added_tables': len(tables.get('added', [])),
            'removed_tables': len(tables.get('removed', [])),
            'modified_tables': len(tables.get('modified', []))
        }
        
        # 表差异列表
        table_diffs = []
        type_labels = {'added': '新增', 'removed': '删除', 'modified': '修改'}
        
        for diff_type in ['added', 'removed', 'modified']:
            for item in tables.get(diff_type, []):
                table_diffs.append({
                    'type': diff_type,
                    'type_label': type_labels[diff_type],
                    'name': item.get('name'),
                    'differences': item.get('differences', [])
                })
        
        # 字段差异
        column_diffs = {}
        for table_name, diffs in columns.items():
            column_diffs[table_name] = {
                'added': [c.get('name') for c in diffs.get('added', [])],
                'removed': [c.get('name') for c in diffs.get('removed', [])],
                'modified': [c.get('name') for c in diffs.get('modified', [])]
            }
        
        return {
            'generated_at': self.generated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'source_name': self.task_config.get('source', {}).get('database', '未知'),
            'target_name': self.task_config.get('target', {}).get('database', '未知'),
            'summary': summary,
            'table_diffs': table_diffs,
            'column_diffs': column_diffs
        }
    
    def _create_summary_sheet(self, ws):
        """创建摘要工作表"""
        # 标题
        ws['A1'] = '数据库比对报告'
        ws['A1'].font = Font(size=18, bold=True)
        ws.merge_cells('A1:D1')
        
        # 元信息
        ws['A3'] = '生成时间'
        ws['B3'] = self.generated_at.strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = '源数据库'
        ws['B4'] = self.task_config.get('source', {}).get('database', '未知')
        ws['A5'] = '目标数据库'
        ws['B5'] = self.task_config.get('target', {}).get('database', '未知')
        
        # 统计信息
        ws['A7'] = '统计摘要'
        ws['A7'].font = Font(size=14, bold=True)
        
        tables = self.result.get('tables', {})
        
        headers = ['类型', '数量', '说明']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6F7FF', end_color='E6F7FF', fill_type='solid')
        
        data = [
            ['新增表', len(tables.get('added', [])), '源库存在但目标库不存在的表'],
            ['删除表', len(tables.get('removed', [])), '目标库存在但源库不存在的表'],
            ['修改表', len(tables.get('modified', [])), '表属性存在差异的表'],
        ]
        
        for row_idx, row_data in enumerate(data, 9):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
    
    def _create_table_diff_sheet(self, ws):
        """创建表结构差异工作表"""
        headers = ['类型', '表名', '差异字段', '源库值', '目标库值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6F7FF', end_color='E6F7FF', fill_type='solid')
        
        tables = self.result.get('tables', {})
        row = 2
        
        type_names = {'added': '新增', 'removed': '删除', 'modified': '修改'}
        type_colors = {
            'added': 'C6EFCE',
            'removed': 'FFC7CE',
            'modified': 'FFEB9C'
        }
        
        for diff_type in ['added', 'removed', 'modified']:
            for item in tables.get(diff_type, []):
                differences = item.get('differences', [])
                
                if differences:
                    for diff in differences:
                        ws.cell(row=row, column=1, value=type_names[diff_type])
                        ws.cell(row=row, column=2, value=item.get('name'))
                        ws.cell(row=row, column=3, value=diff.get('field'))
                        ws.cell(row=row, column=4, value=str(diff.get('source', '')))
                        ws.cell(row=row, column=5, value=str(diff.get('target', '')))
                        
                        # 设置背景色
                        for col in range(1, 6):
                            ws.cell(row=row, column=col).fill = PatternFill(
                                start_color=type_colors[diff_type],
                                end_color=type_colors[diff_type],
                                fill_type='solid'
                            )
                        
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=type_names[diff_type])
                    ws.cell(row=row, column=2, value=item.get('name'))
                    
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=type_colors[diff_type],
                            end_color=type_colors[diff_type],
                            fill_type='solid'
                        )
                    
                    row += 1
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
    
    def _create_column_diff_sheet(self, ws):
        """创建字段差异工作表"""
        headers = ['表名', '字段名', '差异类型', '差异详情']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6F7FF', end_color='E6F7FF', fill_type='solid')
        
        columns = self.result.get('columns', {})
        row = 2
        
        type_names = {'added': '新增', 'removed': '删除', 'modified': '修改'}
        type_colors = {
            'added': 'C6EFCE',
            'removed': 'FFC7CE',
            'modified': 'FFEB9C'
        }
        
        for table_name, diffs in columns.items():
            for diff_type in ['added', 'removed', 'modified']:
                for item in diffs.get(diff_type, []):
                    ws.cell(row=row, column=1, value=table_name)
                    ws.cell(row=row, column=2, value=item.get('name'))
                    ws.cell(row=row, column=3, value=type_names[diff_type])
                    
                    # 差异详情
                    if diff_type == 'modified' and item.get('differences'):
                        details = '; '.join([
                            f"{d.get('field')}: {d.get('source')} → {d.get('target')}"
                            for d in item.get('differences', [])
                        ])
                    else:
                        details = ''
                    
                    ws.cell(row=row, column=4, value=details)
                    
                    # 设置背景色
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=type_colors[diff_type],
                            end_color=type_colors[diff_type],
                            fill_type='solid'
                        )
                    
                    row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
